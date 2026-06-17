#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Handler xử lý học phần từ hệ thống HUTECH.

Cú pháp callback:
    namhoc_<ma_hoc_ky>                    - chọn học kỳ
    hocphan_<ma_hoc_ky>|<key_check>       - xem chi tiết học phần (embed học kỳ để tra cứu đúng)
    hocphan_back                          - quay lại
    danhsach_<key>                        - danh sách sinh viên của lớp
    diemdanh_lop_hoc_phan_<key>           - điểm danh của lớp
"""

import asyncio
import io
import json
import logging
import unicodedata
from datetime import datetime
from typing import Any, Dict, List, Optional

import aiohttp
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from config.config import Config
from utils.button_style import make_inline_button, build_inline_keyboard
from utils.rich_message import (
    escape_html,
    section_heading,
    h1,
    h2,
    p,
    p_bold,
    code,
    hr,
    footer_updated_at,
    join_blocks,
    kv_line,
    table,
)
from utils.telegram_api import TelegramAPI, TelegramAPIError

logger = logging.getLogger(__name__)


class HocPhanHandler:
    def __init__(self, db_manager, cache_manager, telegram_api: Optional[TelegramAPI] = None):
        self.db_manager = db_manager
        self.cache_manager = cache_manager
        self.config = Config()
        self.telegram = telegram_api or TelegramAPI(self.config)

    # ==================== Command ====================

    async def cmd_hocphan(self, chat_id: int, user_id: int, reply_to_message_id: Optional[int]) -> None:
        if not await self.db_manager.is_user_logged_in(user_id):
            await self.telegram.send_message(
                chat_id=chat_id,
                text="Bạn chưa đăng nhập. Vui lòng /dangnhap để đăng nhập.",
                reply_to_message_id=reply_to_message_id,
            )
            return
        result = await self.handle_hoc_phan(user_id)
        if result["success"]:
            await self.telegram.send_rich_message(
                chat_id=chat_id,
                html=self._format_nam_hoc(result["data"]),
                reply_markup=self._build_nam_hoc_keyboard(self.get_nam_hoc_hoc_ky_list(result["data"])),
                reply_to_message_id=reply_to_message_id,
            )
        else:
            await self.telegram.send_message(
                chat_id=chat_id,
                text=result["message"],
                parse_mode="HTML",
                reply_to_message_id=reply_to_message_id,
            )

    # ==================== Callback router ====================

    async def cb_route(self, callback_id: str, chat_id: int, message_id: int,
                       user_id: int, callback_data: str) -> None:
        if callback_data.startswith("namhoc_"):
            await self._cb_namhoc(callback_id, chat_id, message_id, user_id, callback_data)
        elif callback_data.startswith("hocphan_"):
            await self._cb_hocphan(callback_id, chat_id, message_id, user_id, callback_data)
        elif callback_data.startswith("danhsach_"):
            await self._cb_danhsach(callback_id, chat_id, message_id, user_id, callback_data)
        elif callback_data.startswith("diemdanh_lop_hoc_phan_"):
            await self._cb_diemdanh_lop(callback_id, chat_id, message_id, user_id, callback_data)

    # ==================== Callback implementations ====================

    async def _cb_namhoc(self, callback_id: str, chat_id: int, message_id: int,
                        user_id: int, callback_data: str) -> None:
        nam_hoc_key = callback_data[len("namhoc_"):]
        await self.telegram.answer_callback_query(callback_id, text="Đang tìm kiếm học phần...")

        result = await self.handle_hoc_phan(user_id)
        if not result["success"]:
            try:
                await self.telegram.edit_message_text_plain(
                    chat_id=chat_id, message_id=message_id, text=result["message"], parse_mode="HTML"
                )
            except TelegramAPIError:
                pass
            return

        search_result = await self.handle_search_hoc_phan(user_id, [nam_hoc_key])
        if not search_result["success"]:
            try:
                await self.telegram.edit_message_text_plain(
                    chat_id=chat_id,
                    message_id=message_id,
                    text=search_result["message"],
                    reply_markup=build_inline_keyboard([[
                        make_inline_button("Quay lại", "hocphan_back", tone="neutral")
                    ]]),
                    parse_mode="HTML",
                )
            except TelegramAPIError:
                pass
            return

        try:
            await self.telegram.edit_message_text_rich(
                chat_id=chat_id,
                message_id=message_id,
                html=self._format_search(search_result["data"]),
                reply_markup=self._build_hocphan_keyboard(self.get_hoc_phan_list(search_result["data"])),
            )
        except TelegramAPIError as e:
            if "message is not modified" not in e.description.lower():
                raise

    async def _cb_hocphan(self, callback_id: str, chat_id: int, message_id: int,
                         user_id: int, callback_data: str) -> None:
        if callback_data == "hocphan_back":
            await self.telegram.answer_callback_query(callback_id)
            result = await self.handle_hoc_phan(user_id)
            if result["success"]:
                try:
                    await self.telegram.edit_message_text_rich(
                        chat_id=chat_id,
                        message_id=message_id,
                        html=self._format_nam_hoc(result["data"]),
                        reply_markup=self._build_nam_hoc_keyboard(self.get_nam_hoc_hoc_ky_list(result["data"])),
                    )
                except TelegramAPIError as e:
                    if "message is not modified" not in e.description.lower():
                        raise
            return

        key_lop_hoc_phan = callback_data[len("hocphan_"):]
        await self.telegram.answer_callback_query(callback_id)

        # Tách ma_hoc_ky và key_check. Định dạng callback: hocphan_<ma_hoc_ky>|<key_check>
        # Nếu callback cũ không có '|' thì fallback về logic tìm theo học kỳ mới nhất.
        if "|" in key_lop_hoc_phan:
            nam_hoc_key, key_lop_hoc_phan = key_lop_hoc_phan.split("|", 1)
        else:
            result = await self.handle_hoc_phan(user_id)
            if not result["success"]:
                return
            list_nhhk = self.get_nam_hoc_hoc_ky_list(result["data"])
            nam_hoc_key = list_nhhk[0]["key"] if list_nhhk else None
            if not nam_hoc_key:
                await self._safe_edit(
                    chat_id, message_id, "Không có năm học - học kỳ nào để tìm kiếm."
                )
                return

        search_result = await self.handle_search_hoc_phan(user_id, [nam_hoc_key])
        if not search_result["success"]:
            await self._safe_edit(chat_id, message_id, search_result["message"], parse_mode="HTML")
            return
        hoc_phan_list = search_result["data"].get("hoc_phan_list", [])
        selected = next(
            (hp for hp in hoc_phan_list if hp.get("key_check") == key_lop_hoc_phan), None
        )
        if not selected:
            await self._safe_edit(chat_id, message_id, "Không tìm thấy học phần được chọn.")
            return
        try:
            await self.telegram.edit_message_text_rich(
                chat_id=chat_id,
                message_id=message_id,
                html=self._format_hoc_phan_detail(selected),
                reply_markup=build_inline_keyboard([
                    [
                        make_inline_button("Danh sách sinh viên", f"danhsach_{key_lop_hoc_phan}", tone="primary", emoji="📋"),
                        make_inline_button("Điểm danh", f"diemdanh_lop_hoc_phan_{key_lop_hoc_phan}", tone="success", emoji="📝"),
                    ],
                    [make_inline_button("Quay lại", "hocphan_back", tone="neutral")],
                ]),
            )
        except TelegramAPIError as e:
            if "message is not modified" not in e.description.lower():
                raise

    async def _cb_danhsach(self, callback_id: str, chat_id: int, message_id: int,
                          user_id: int, callback_data: str) -> None:
        key = callback_data[len("danhsach_"):]
        await self.telegram.answer_callback_query(callback_id, text="Đang tải danh sách sinh viên...")
        result = await self.handle_danh_sach_sinh_vien(user_id, key)
        if not result["success"]:
            await self._safe_edit(chat_id, message_id, result["message"], parse_mode="HTML")
            return
        try:
            file_bytes = await asyncio.to_thread(self.generate_danh_sach_sinh_vien_xlsx, result["data"])
            await self.telegram.send_document(
                chat_id=chat_id,
                file=file_bytes,
                filename=f"danh_sach_sinh_vien_{key}.xlsx",
                caption="📋 Danh sách sinh viên lớp học phần",
            )
        except Exception as e:
            logger.error("Lỗi tạo/gửi file Excel: %s", e)
            await self.telegram.send_message(
                chat_id=chat_id, text=f"Lỗi tạo file Excel: {str(e)}"
            )
        # Xoá menu cũ
        await self.telegram.delete_message(chat_id, message_id)

    async def _cb_diemdanh_lop(self, callback_id: str, chat_id: int, message_id: int,
                              user_id: int, callback_data: str) -> None:
        key = callback_data[len("diemdanh_lop_hoc_phan_"):]
        await self.telegram.answer_callback_query(callback_id, text="Đang tải lịch sử điểm danh...")
        result = await self.handle_diem_danh(user_id, key)
        if result["success"]:
            try:
                await self.telegram.edit_message_text_rich(
                    chat_id=chat_id,
                    message_id=message_id,
                    html=self._format_diem_danh(result["data"]),
                    reply_markup=build_inline_keyboard([
                        [make_inline_button("Quay lại", "hocphan_back", tone="neutral")]
                    ]),
                )
            except TelegramAPIError as e:
                if "message is not modified" not in e.description.lower():
                    raise
        else:
            await self._safe_edit(chat_id, message_id, result["message"], parse_mode="HTML")

    # ==================== Data layer ====================

    async def handle_hoc_phan(self, telegram_user_id: int) -> Dict[str, Any]:
        try:
            cache_key = f"nam_hoc_hoc_ky:{telegram_user_id}"
            cached = await self.cache_manager.get(cache_key)
            if cached:
                processed = self._process_nam_hoc_hoc_ky_data(cached.get("data"))
                processed["timestamp"] = cached.get("timestamp")
                return {"success": True, "message": "OK", "data": processed}

            token = await self._get_user_token(telegram_user_id)
            if not token:
                return {"success": False, "message": "Bạn chưa đăng nhập. Vui lòng /dangnhap.", "data": None}
            response = await self._call_nam_hoc_hoc_ky_api(token)
            if response and isinstance(response, list):
                await self.cache_manager.set(cache_key, response, ttl=86400)
                processed = self._process_nam_hoc_hoc_ky_data(response)
                processed["timestamp"] = datetime.utcnow().isoformat()
                return {"success": True, "message": "OK", "data": processed}
            return {"success": False, "message": "🚫 Lỗi: Không thể lấy danh sách năm học - học kỳ.", "data": response}
        except Exception as e:
            logger.error("Học phần error for user %s: %s", telegram_user_id, e)
            return {"success": False, "message": f"🚫 Lỗi: {str(e)}", "data": None}

    async def handle_search_hoc_phan(self, telegram_user_id: int, nam_hoc_hoc_ky_list: List[str]) -> Dict[str, Any]:
        try:
            cache_key = f"search_hoc_phan:{telegram_user_id}:{':'.join(sorted(nam_hoc_hoc_ky_list))}"
            cached = await self.cache_manager.get(cache_key)
            if cached:
                return {"success": True, "message": "OK (cache)", "data": self._process_search_hoc_phan_data(cached.get("data", []))}

            token = await self._get_user_token(telegram_user_id)
            if not token:
                return {"success": False, "message": "Bạn chưa đăng nhập. Vui lòng /dangnhap.", "data": None}

            response = await self._call_search_hoc_phan_api(token, nam_hoc_hoc_ky_list)
            if not (response and isinstance(response, list)):
                logger.warning("Search học phần thất bại lần 1, retry với renew=true")
                response = await self._call_search_hoc_phan_api(token, nam_hoc_hoc_ky_list, use_renew=True)
            if response and isinstance(response, list):
                await self.cache_manager.set(cache_key, response, ttl=3600)
                processed = self._process_search_hoc_phan_data(response)
                processed["timestamp"] = datetime.utcnow().isoformat()
                return {"success": True, "message": "OK", "data": processed}
            return {"success": False, "message": "🚫 Lỗi: Không thể tìm kiếm học phần.", "data": response}
        except Exception as e:
            logger.error("Search học phần error for user %s: %s", telegram_user_id, e)
            return {"success": False, "message": f"🚫 Lỗi: {str(e)}", "data": None}

    async def handle_diem_danh(self, telegram_user_id: int, key_lop_hoc_phan: str) -> Dict[str, Any]:
        try:
            cache_key = f"diem_danh:{telegram_user_id}:{key_lop_hoc_phan}"
            cached = await self.cache_manager.get(cache_key)
            if cached:
                return {"success": True, "message": "OK (cache)", "data": self._process_diem_danh_data(cached.get("data", []))}
            token = await self._get_user_token(telegram_user_id)
            if not token:
                return {"success": False, "message": "Bạn chưa đăng nhập. Vui lòng /dangnhap.", "data": None}
            response = await self._call_diem_danh_api(token, key_lop_hoc_phan)
            if response and isinstance(response, dict) and "result" in response:
                await self.cache_manager.set(cache_key, response["result"], ttl=3600)
                processed = self._process_diem_danh_data(response["result"])
                processed["timestamp"] = datetime.utcnow().isoformat()
                return {"success": True, "message": "OK", "data": processed}
            error_message = "Danh sách điểm danh chưa được cập nhật"
            if response and response.get("error"):
                try:
                    api_err = json.loads(response.get("message", "{}"))
                    extracted = api_err.get("reasons", {}).get("message") or api_err.get("errorMessage")
                    if extracted:
                        error_message = extracted.split(" - ", 1)[-1]
                except (json.JSONDecodeError, AttributeError):
                    if isinstance(response.get("message"), str):
                        error_message = response["message"]
            return {"success": False, "message": f"🚫 Lỗi: {error_message}", "data": response}
        except Exception as e:
            logger.error("Điểm danh error for user %s: %s", telegram_user_id, e)
            return {"success": False, "message": f"🚫 Lỗi: {str(e)}", "data": None}

    async def handle_danh_sach_sinh_vien(self, telegram_user_id: int, key_lop_hoc_phan: str) -> Dict[str, Any]:
        try:
            token = await self._get_user_token(telegram_user_id)
            if not token:
                return {"success": False, "message": "Bạn chưa đăng nhập. Vui lòng /dangnhap.", "data": None}
            response = await self._call_danh_sach_sinh_vien_api(token, key_lop_hoc_phan)
            if response and isinstance(response, dict):
                return {"success": True, "message": "OK", "data": self._process_danh_sach_sinh_vien_data(response)}
            return {"success": False, "message": "🚫 Lỗi: Không thể lấy danh sách sinh viên.", "data": response}
        except Exception as e:
            logger.error("Danh sách sinh viên error for user %s: %s", telegram_user_id, e)
            return {"success": False, "message": f"🚫 Lỗi: {str(e)}", "data": None}

    # ==================== API calls ====================

    async def _call_nam_hoc_hoc_ky_api(self, token: str) -> Optional[Any]:
        try:
            url = f"{self.config.HUTECH_API_BASE_URL}{self.config.HUTECH_HOC_PHAN_NAM_HOC_HOC_KY_ENDPOINT}"
            headers = self.config.HUTECH_MOBILE_HEADERS.copy()
            headers["authorization"] = f"JWT {token}"
            async with aiohttp.ClientSession() as session:
                async with session.get(url, headers=headers) as response:
                    if response.status == 200:
                        return await response.json()
                    return {"error": True, "status_code": response.status, "message": await response.text()}
        except Exception as e:
            logger.error("Lỗi gọi nam_hoc_hoc_ky API: %s", e)
            return {"error": True, "message": f"Lỗi: {str(e)}"}

    async def _call_search_hoc_phan_api(self, token: str, nam_hoc_hoc_ky_list: List[str], use_renew: bool = False) -> Optional[Any]:
        try:
            url = f"{self.config.HUTECH_API_BASE_URL}{self.config.HUTECH_HOC_PHAN_SEARCH_ENDPOINT}"
            headers = self.config.HUTECH_MOBILE_HEADERS.copy()
            headers["authorization"] = f"JWT {token}"
            body = {"nam_hoc_hoc_ky": nam_hoc_hoc_ky_list}
            if use_renew:
                body = {"renew": True, "nam_hoc_hoc_ky": nam_hoc_hoc_ky_list}
            async with aiohttp.ClientSession() as session:
                async with session.post(url, headers=headers, json=body) as response:
                    if response.status == 200:
                        return await response.json()
                    return {"error": True, "status_code": response.status, "message": await response.text()}
        except Exception as e:
            logger.error("Lỗi gọi search hoc_phan API: %s", e)
            return {"error": True, "message": f"Lỗi: {str(e)}"}

    async def _call_diem_danh_api(self, token: str, key_lop_hoc_phan: str) -> Optional[Any]:
        try:
            url = f"{self.config.HUTECH_API_BASE_URL}{self.config.HUTECH_HOC_PHAN_DIEM_DANH_ENDPOINT}"
            headers = self.config.HUTECH_MOBILE_HEADERS.copy()
            headers["authorization"] = f"JWT {token}"
            params = {"key_lop_hoc_phan": key_lop_hoc_phan}
            async with aiohttp.ClientSession() as session:
                async with session.get(url, headers=headers, params=params) as response:
                    if response.status == 200:
                        return await response.json()
                    return {"error": True, "status_code": response.status, "message": await response.text()}
        except Exception as e:
            logger.error("Lỗi gọi diem_danh API: %s", e, exc_info=True)
            return {"error": True, "message": f"Lỗi: {str(e)}"}

    async def _call_danh_sach_sinh_vien_api(self, token: str, key_lop_hoc_phan: str) -> Optional[Any]:
        try:
            url = f"{self.config.HUTECH_API_BASE_URL}{self.config.HUTECH_HOC_PHAN_DANH_SACH_SINH_VIEN_ENDPOINT}"
            headers = self.config.HUTECH_MOBILE_HEADERS.copy()
            headers["authorization"] = f"JWT {token}"
            params = {"key_lop_hoc_phan": key_lop_hoc_phan}
            async with aiohttp.ClientSession() as session:
                async with session.get(url, headers=headers, params=params) as response:
                    if response.status == 200:
                        return await response.json()
                    return {"error": True, "status_code": response.status, "message": await response.text()}
        except Exception as e:
            logger.error("Lỗi gọi danh_sach_sinh_vien API: %s", e)
            return {"error": True, "message": f"Lỗi: {str(e)}"}

    async def _get_user_token(self, telegram_user_id: int) -> Optional[str]:
        try:
            response_data = await self.db_manager.get_user_login_response(telegram_user_id)
            if not response_data:
                return None
            old_login_info = response_data.get("old_login_info")
            if isinstance(old_login_info, dict) and old_login_info.get("token"):
                return old_login_info["token"]
            return response_data.get("token")
        except Exception as e:
            logger.error("Error getting token for user %s: %s", telegram_user_id, e)
            return None

    # ==================== Process & format ====================

    @staticmethod
    def _format_hoc_ky_label(hoc_ky: Any) -> str:
        hoc_ky_str = str(hoc_ky).strip()
        if not hoc_ky_str or hoc_ky_str == "N/A":
            return "N/A"
        normalized = hoc_ky_str.lstrip("0") or hoc_ky_str
        mapping = {
            "1": "HK1", "2": "HK phụ HK1", "3": "HK2",
            "4": "HK phụ HK2", "5": "HK3",
        }
        return mapping.get(normalized, hoc_ky_str)

    @staticmethod
    def _format_hoc_ky_excel_label(hoc_ky: Any) -> str:
        hoc_ky_str = str(hoc_ky).strip()
        if not hoc_ky_str or hoc_ky_str == "N/A":
            return "N/A"
        normalized = hoc_ky_str.lstrip("0") or hoc_ky_str
        mapping = {"1": "1", "2": "phụ HK1", "3": "2", "4": "phụ HK2", "5": "3"}
        return mapping.get(normalized, hoc_ky_str)

    @staticmethod
    def _process_nam_hoc_hoc_ky_data(data: List[Dict[str, Any]]) -> Dict[str, Any]:
        try:
            filtered = [item for item in data if int(item.get("ma_hoc_ky", "0")[-1]) % 2 != 0]
            sorted_data = sorted(filtered, key=lambda x: x.get("ma_hoc_ky", ""), reverse=True)
            return {"nam_hoc_hoc_ky_list": sorted_data}
        except Exception as e:
            logger.error("Error processing năm học - học kỳ data: %s", e)
            return {"nam_hoc_hoc_ky_list": []}

    @staticmethod
    def _process_search_hoc_phan_data(data: List[Dict[str, Any]]) -> Dict[str, Any]:
        try:
            sorted_data = sorted(data, key=lambda x: (
                x.get("json_thong_tin", {}).get("nam_hoc", ""),
                x.get("json_thong_tin", {}).get("hoc_ky", ""),
                x.get("json_thong_tin", {}).get("ten_mon_hoc", ""),
            ), reverse=True)
            return {"hoc_phan_list": sorted_data}
        except Exception as e:
            logger.error("Error processing search học phần data: %s", e)
            return {"hoc_phan_list": []}

    @staticmethod
    def _process_diem_danh_data(data: List[Dict[str, Any]]) -> Dict[str, Any]:
        def parse_date(s):
            try:
                return datetime.strptime(s, "%d/%m/%Y")
            except (ValueError, TypeError):
                return datetime.max
        try:
            sorted_data = sorted(
                data, key=lambda x: parse_date(x.get("lich_trinh", {}).get("ngay_hoc", ""))
            )
            return {"diem_danh_list": sorted_data}
        except Exception as e:
            logger.error("Error processing điểm danh data: %s", e, exc_info=True)
            return {"diem_danh_list": []}

    def _process_danh_sach_sinh_vien_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        try:
            lop_info = data.get("lop", {})
            json_member = lop_info.get("json_member", {})
            sinh_vien_list = []
            for mssv, info in json_member.items():
                ho_ten = info.get("ho_ten", "")
                parts = ho_ten.split()
                if len(parts) > 1:
                    ho = " ".join(parts[:-1])
                    ten = parts[-1]
                else:
                    ho = ""
                    ten = ho_ten
                sinh_vien_list.append({
                    "mssv": mssv, "ho": ho, "ten": ten,
                    "lop": info.get("lop", ""), "ho_ten_day_du": ho_ten,
                })
            sinh_vien_list.sort(key=lambda s: (
                self._build_vietnamese_sort_key(s["ten"]),
                self._build_vietnamese_sort_key(s["ho"]),
                s["mssv"],
            ))
            return {"lop_info": lop_info, "sinh_vien_list": sinh_vien_list}
        except Exception as e:
            logger.error("Error processing danh sách sinh viên data: %s", e)
            return {"lop_info": {}, "sinh_vien_list": []}

    # ==================== Formatting ====================

    def _format_nam_hoc(self, data: Dict[str, Any]) -> str:
        items = data.get("nam_hoc_hoc_ky_list", [])
        if not items:
            return join_blocks([
                section_heading("📚", "Học Phần"),
                p("Không có dữ liệu năm học - học kỳ."),
            ])
        rows: List[List[str]] = []
        for i, item in enumerate(items, 1):
            ma = item.get("ma_hoc_ky", "N/A")
            ten = item.get("ten_hoc_ky", "N/A")
            rows.append([str(i), ten, ma])
        blocks: List[str] = [
            section_heading("📚", "Danh Sách Năm Học - Học Kỳ"),
            p("Chọn một học kỳ bên dưới để tìm kiếm học phần."),
            table(
                ["#", "Tên học kỳ", "Mã học kỳ"],
                rows,
                bordered=True,
                striped=True,
            ),
        ]
        ts = footer_updated_at(data.get("timestamp"))
        if ts:
            blocks.append(ts)
        return join_blocks(blocks)

    def _format_search(self, data: Dict[str, Any]) -> str:
        items = data.get("hoc_phan_list", [])
        if not items:
            return join_blocks([
                section_heading("📚", "Kết Quả Tìm Kiếm"),
                p("Không có học phần nào được tìm thấy."),
            ])
        rows: List[List[str]] = []
        for i, item in enumerate(items, 1):
            tt = item.get("json_thong_tin", {})
            ten = tt.get("ten_mon_hoc", "N/A")
            ma = tt.get("ma_mon_hoc", "N/A")
            nam_hoc = tt.get("nam_hoc", "N/A")
            hoc_ky = tt.get("hoc_ky", "N/A")
            nhom = tt.get("nhom_hoc", "N/A")
            tc = tt.get("so_tc", "N/A")
            hoc_ky_label = self._format_hoc_ky_label(hoc_ky)
            rows.append([
                str(i), ten, ma, f"{nam_hoc} - {hoc_ky_label}", str(nhom), str(tc),
            ])
        blocks: List[str] = [
            section_heading("📚", "Kết Quả Tìm Kiếm Học Phần"),
            table(
                ["#", "Tên môn học", "Mã HP", "Học kỳ", "Nhóm", "Số TC"],
                rows,
                bordered=True,
                striped=True,
            ),
        ]
        ts = footer_updated_at(data.get("timestamp"))
        if ts:
            blocks.append(ts)
        return join_blocks(blocks)

    def _format_hoc_phan_detail(self, hoc_phan: Dict[str, Any]) -> str:
        tt = hoc_phan.get("json_thong_tin", {})
        ten = tt.get("ten_mon_hoc", "N/A")
        ma = tt.get("ma_mon_hoc", "N/A")
        nam_hoc = tt.get("nam_hoc", "N/A")
        hoc_ky = tt.get("hoc_ky", "N/A")
        nhom = tt.get("nhom_hoc", "N/A")
        tc = tt.get("so_tc", "N/A")
        nhom_th = tt.get("nhom_thuc_hanh", "")
        hoc_ky_label = self._format_hoc_ky_label(hoc_ky)
        info_rows: List[List[str]] = [
            ["Mã HP", ma],
            ["Học kỳ", f"{nam_hoc} - {hoc_ky_label}"],
            ["Nhóm", str(nhom)],
            ["Số TC", str(tc)],
        ]
        if nhom_th:
            info_rows.append(["Nhóm TH", str(nhom_th)])
        blocks: List[str] = [
            section_heading("📚", "Chi Tiết Học Phần"),
            p_bold(ten),
            table(
                ["Chỉ số", "Giá trị"],
                info_rows,
                bordered=True,
                striped=True,
            ),
        ]
        ts = footer_updated_at(hoc_phan.get("timestamp"))
        if ts:
            blocks.append(ts)
        return join_blocks(blocks)

    def _format_diem_danh(self, data: Dict[str, Any]) -> str:
        items = data.get("diem_danh_list", [])
        if not items:
            return join_blocks([
                section_heading("📝", "Lịch Sử Điểm Danh"),
                p("Không có dữ liệu điểm danh."),
            ])
        total = len(items)
        present = sum(1 for it in items if (it.get("diem_danh") or {}).get("ket_qua") == "co_mat")
        absent = sum(1 for it in items if (it.get("diem_danh") or {}).get("ket_qua") in {"vang_mat", "vang"})

        summary_rows: List[List[str]] = [
            ["✅ Có mặt", f"{present}/{total}"],
            ["❌ Vắng mặt", f"{absent}/{total}"],
        ]
        rows: List[List[str]] = []
        for it in items:
            if not it:
                continue
            lich_trinh = it.get("lich_trinh", {})
            dd = it.get("diem_danh") or {}
            ngay = lich_trinh.get("ngay_hoc", "—")
            bd = lich_trinh.get("gio_bat_dau", "—")
            kt = lich_trinh.get("gio_ket_thuc", "—")
            phong = lich_trinh.get("ma_phong", "—")
            ket_qua = dd.get("ket_qua", "chua_diem_danh")
            if ket_qua == "co_mat":
                status = "✅ Có mặt"
            elif ket_qua in {"vang_mat", "vang"}:
                status = "❌ Vắng mặt"
            else:
                status = "❔ Chưa điểm danh"
            chi_tiet = (lich_trinh.get("diem_danh") or {}).get("chi_tiet", [])
            extra = ""
            if chi_tiet:
                qr = (chi_tiet[0].get("diem_danh_qr_code") or {}).get("data", {})
                thoi_gian = qr.get("time")
                vi_tri = (qr.get("location") or {}).get("display_name")
                bits = []
                if thoi_gian: bits.append(f"TG: {thoi_gian}")
                if vi_tri: bits.append(f"VT: {vi_tri}")
                extra = "\n".join(bits)
            rows.append([
                ngay, f"{bd} - {kt}", phong, status, extra or "—",
            ])

        blocks: List[str] = [
            section_heading("📝", "Lịch Sử Điểm Danh"),
            h2("Tổng quan"),
            table(
                ["Chỉ số", "Giá trị"],
                summary_rows,
                bordered=True,
                striped=True,
            ),
            hr(),
            h2("Chi tiết"),
            table(
                ["Ngày học", "Giờ", "Phòng", "Trạng thái", "Ghi chú"],
                rows,
                bordered=True,
                striped=True,
            ),
        ]
        ts = footer_updated_at(data.get("timestamp"))
        if ts:
            blocks.append(ts)
        return join_blocks(blocks)

    # ==================== Timestamp helper ====================
    # (đã chuyển sang dùng utils.rich_message.footer_updated_at)

    # ==================== Keyboard ====================

    def get_nam_hoc_hoc_ky_list(self, data: Dict[str, Any]) -> List[Dict[str, str]]:
        result: List[Dict[str, str]] = []
        for i, item in enumerate(data.get("nam_hoc_hoc_ky_list", [])):
            result.append({
                "key": item.get("ma_hoc_ky", ""),
                "name": item.get("ten_hoc_ky", ""),
                "display": str(i + 1),
            })
        return result

    def get_hoc_phan_list(self, data: Dict[str, Any]) -> List[Dict[str, str]]:
        result: List[Dict[str, str]] = []
        for i, item in enumerate(data.get("hoc_phan_list", [])):
            tt = item.get("json_thong_tin", {})
            ten = tt.get("ten_mon_hoc", "")
            ma = tt.get("ma_mon_hoc", "")
            nam_hoc = tt.get("nam_hoc", "")
            hoc_ky = tt.get("hoc_ky", "")
            nhom = tt.get("nhom_hoc", "")
            hoc_ky_label = self._format_hoc_ky_label(hoc_ky)
            display = f"{ten} ({ma})"
            if len(display) > 40:
                display = display[:37] + "..."
            ma_hoc_ky = item.get("ma_hoc_ky") or f"{nam_hoc}{hoc_ky}"
            result.append({
                "key": item.get("key_check", ""),
                "name": display,
                "full_name": f"{ten} ({ma}) - {nam_hoc} - {hoc_ky_label} - NH{nhom}",
                "display": str(i + 1),
                "ma_hoc_ky": ma_hoc_ky,
            })
        return result

    def _build_nam_hoc_keyboard(self, items: List[Dict[str, str]]) -> Dict[str, Any]:
        rows: List[List[Dict[str, Any]]] = []
        row: List[Dict[str, Any]] = []
        for i, it in enumerate(items):
            row.append(make_inline_button(it["name"], f"namhoc_{it['key']}", tone=None))
            if len(row) == 3 or i == len(items) - 1:
                rows.append(row)
                row = []
        return build_inline_keyboard(rows)

    def _build_hocphan_keyboard(self, items: List[Dict[str, str]]) -> Dict[str, Any]:
        rows: List[List[Dict[str, Any]]] = []
        row: List[Dict[str, Any]] = []
        for i, it in enumerate(items):
            cb_data = f"hocphan_{it['ma_hoc_ky']}|{it['key']}" if it.get("ma_hoc_ky") else f"hocphan_{it['key']}"
            row.append(make_inline_button(it["name"], cb_data, tone=None))
            if len(row) == 2 or i == len(items) - 1:
                rows.append(row)
                row = []
        rows.append([make_inline_button("Quay lại", "hocphan_back", tone="neutral")])
        return build_inline_keyboard(rows)

    # ==================== Excel ====================

    def generate_danh_sach_sinh_vien_xlsx(self, data: Dict[str, Any]) -> bytes:
        try:
            lop_info = data.get("lop_info", {})
            sinh_vien_list = data.get("sinh_vien_list", [])
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Danh sách sinh viên"
            title_font = Font(name='Arial', size=14, bold=True)
            header_font = Font(name='Arial', size=12, bold=True)
            cell_font = Font(name='Arial', size=11)
            header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center')
            cell_align = Alignment(horizontal='left', vertical='center')
            stt_align = Alignment(horizontal='center', vertical='center')

            tt = lop_info.get("json_thong_tin", {})
            ten_mon = tt.get("ten_mon_hoc", "")
            ma_mon = tt.get("ma_mon_hoc", "")
            nam_hoc = tt.get("nam_hoc", "")
            hoc_ky = tt.get("hoc_ky", "")
            nhom = tt.get("nhom_hoc", "")
            hoc_ky_label = self._format_hoc_ky_excel_label(hoc_ky)

            ws.merge_cells('A1:E1')
            ws['A1'] = "DANH SÁCH SINH VIÊN LỚP HỌC PHẦN"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells('A2:E2')
            ws['A2'] = f"{ten_mon} ({ma_mon})"
            ws['A2'].font = header_font
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells('A3:E3')
            ws['A3'] = f"Năm học: {nam_hoc} - Học kỳ: {hoc_ky_label} - Nhóm học: {nhom}"
            ws['A3'].font = cell_font
            ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

            headers = ['STT', 'MSSV', 'Họ', 'Tên', 'Lớp']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            for row_num, sv in enumerate(sinh_vien_list, 6):
                ws.cell(row=row_num, column=1, value=row_num - 5).font = cell_font
                ws.cell(row=row_num, column=1).alignment = stt_align
                ws.cell(row=row_num, column=2, value=sv["mssv"]).font = cell_font
                ws.cell(row=row_num, column=2).alignment = cell_align
                ws.cell(row=row_num, column=3, value=sv["ho"]).font = cell_font
                ws.cell(row=row_num, column=3).alignment = cell_align
                ws.cell(row=row_num, column=4, value=sv["ten"]).font = cell_font
                ws.cell(row=row_num, column=4).alignment = cell_align
                ws.cell(row=row_num, column=5, value=sv["lop"]).font = cell_font
                ws.cell(row=row_num, column=5).alignment = cell_align
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()
        except Exception as e:
            logger.error("Error generating danh sách sinh viên XLSX: %s", e)
            raise

    # ==================== Vietnamese sort ====================

    @staticmethod
    def _get_vietnamese_char_sort_key(char: str) -> tuple:
        alphabet_order = {
            "a": 0, "ă": 1, "â": 2, "b": 3, "c": 4, "d": 5, "đ": 6,
            "e": 7, "ê": 8, "g": 9, "h": 10, "i": 11, "k": 12, "l": 13,
            "m": 14, "n": 15, "o": 16, "ô": 17, "ơ": 18, "p": 19, "q": 20,
            "r": 21, "s": 22, "t": 23, "u": 24, "ư": 25, "v": 26, "x": 27, "y": 28,
        }
        tone_order = {"": 0, "̀": 1, "̉": 2, "̃": 3, "́": 4, "̣": 5}
        normalized = unicodedata.normalize("NFD", char.casefold())
        if not normalized:
            return (len(alphabet_order), 0, 0)
        if char.casefold() == "đ":
            return (alphabet_order["đ"], 0, ord("đ"))
        base = normalized[0]
        marks = set(normalized[1:])
        if base == "a":
            letter = "ă" if "̆" in marks else "â" if "̂" in marks else "a"
        elif base == "e":
            letter = "ê" if "̂" in marks else "e"
        elif base == "o":
            letter = "ơ" if "̛" in marks else "ô" if "̂" in marks else "o"
        elif base == "u":
            letter = "ư" if "̛" in marks else "u"
        else:
            letter = base
        tone = 0
        for m, rank in tone_order.items():
            if m and m in marks:
                tone = rank
                break
        letter_rank = alphabet_order.get(letter, len(alphabet_order) + ord(base))
        return (letter_rank, tone, ord(base))

    @staticmethod
    def _build_vietnamese_sort_key(value: str) -> tuple:
        normalized = "".join((value or "").split()).casefold()
        return tuple(HocPhanHandler._get_vietnamese_char_sort_key(c) for c in normalized)

    # ==================== Utils ====================

    async def _safe_edit(self, chat_id: int, message_id: int, text: str, **kwargs) -> None:
        try:
            await self.telegram.edit_message_text_plain(
                chat_id=chat_id, message_id=message_id, text=text, **kwargs
            )
        except TelegramAPIError as e:
            if "message is not modified" not in e.description.lower():
                logger.debug("safe_edit failed: %s", e.description)
