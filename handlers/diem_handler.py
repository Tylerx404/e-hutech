#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Handler xử lý điểm từ hệ thống HUTECH.

Lấy bảng điểm theo từng học kỳ, xem chi tiết, xuất Excel.

Sử dụng Rich Message (Bot API 10.1):
- Bảng điểm chi tiết: `<table bordered>` với 10 cột
  (STT, Mã HP, Tên, STC, KT1, KT2, Thi, Điểm 10, Điểm 4, Điểm chữ).
- Menu học kỳ: `<h2>` cho mỗi học kỳ.

Cú pháp callback:
    diem_<hocky_key>         - xem chi tiết học kỳ
    diem_more                - danh sách học kỳ cũ hơn
    diem_back                - quay lại menu chính
    diem_export_all          - xuất Excel toàn bộ
    diem_export_<hocky_key>  - xuất Excel học kỳ
"""

import asyncio
import io
import json
import logging
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

import aiohttp
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from config.config import Config
from utils.button_style import make_inline_button, build_inline_keyboard
from utils.rich_message import (
    escape_html,
    section_heading,
    h2,
    h1,
    hr,
    p,
    p_bold,
    code,
    footer,
    footer_updated_at,
    table,
    join_blocks,
    kv_line,
)
from utils.telegram_api import TelegramAPI, TelegramAPIError

logger = logging.getLogger(__name__)


class DiemHandler:
    """Handler cho `/diem` — xem bảng điểm, xuất Excel."""

    def __init__(self, db_manager, cache_manager, telegram_api: Optional[TelegramAPI] = None):
        self.db_manager = db_manager
        self.cache_manager = cache_manager
        self.config = Config()
        self.telegram = telegram_api or TelegramAPI(self.config)

    # ==================== Command ====================

    async def cmd_diem(self, chat_id: int, user_id: int, reply_to_message_id: Optional[int]) -> None:
        if not await self.db_manager.is_user_logged_in(user_id):
            await self.telegram.send_message(
                chat_id=chat_id,
                text="Bạn chưa đăng nhập. Vui lòng /dangnhap để đăng nhập.",
                reply_to_message_id=reply_to_message_id,
            )
            return
        result = await self.handle_diem(user_id)
        if result["success"]:
            await self.telegram.send_rich_message(
                chat_id=chat_id,
                html=self.format_diem_menu_message(result["data"]),
                reply_markup=self.create_main_diem_keyboard(result["data"]),
                reply_to_message_id=reply_to_message_id,
            )
        else:
            await self.telegram.send_message(
                chat_id=chat_id, text=result["message"], parse_mode="HTML",
                reply_to_message_id=reply_to_message_id,
            )

    # ==================== Callback router ====================

    async def cb_route(self, callback_id: str, chat_id: int, message_id: int,
                       user_id: int, callback_data: str) -> None:
        if not callback_data.startswith("diem_"):
            return
        hocky_key = callback_data[len("diem_"):]

        if hocky_key == "more":
            await self.telegram.answer_callback_query(callback_id, text="Đang tải điểm...")
            result = await self.handle_diem(user_id)
            if not result["success"]:
                await self._safe_edit(chat_id, message_id, result["message"], parse_mode="HTML")
                return
            older = self.get_older_hocky_list(result["data"])
            if not older:
                await self._safe_edit(chat_id, message_id, "Không có học kỳ cũ hơn để hiển thị.")
                return
            rows: List[List[Dict[str, Any]]] = []
            for h in older:
                rows.append([make_inline_button(h["name"], f"diem_{h['key']}", tone=None)])
            rows.append([make_inline_button("Quay lại", "diem_back", tone="neutral")])
            try:
                await self.telegram.edit_message_text_rich(
                    chat_id=chat_id,
                    message_id=message_id,
                    html=self.format_older_hocky_menu_message(result["data"]),
                    reply_markup=build_inline_keyboard(rows),
                )
            except TelegramAPIError as e:
                if "message is not modified" not in e.description.lower():
                    raise
            return

        if hocky_key == "back":
            await self.telegram.answer_callback_query(callback_id, text="Đang tải điểm...")
            result = await self.handle_diem(user_id)
            if not result["success"]:
                await self._safe_edit(chat_id, message_id, result["message"], parse_mode="HTML")
                return
            try:
                await self.telegram.edit_message_text_rich(
                    chat_id=chat_id,
                    message_id=message_id,
                    html=self.format_diem_menu_message(result["data"]),
                    reply_markup=self.create_main_diem_keyboard(result["data"]),
                )
            except TelegramAPIError as e:
                if "message is not modified" not in e.description.lower():
                    raise
            return

        if hocky_key.startswith("export_"):
            export_type = hocky_key[len("export_"):]
            await self.telegram.answer_callback_query(callback_id, text="Đang tạo file Excel...")
            result = await self.handle_diem(user_id)
            if not result["success"]:
                await self._safe_edit(chat_id, message_id, result["message"], parse_mode="HTML")
                return
            try:
                if export_type == "all":
                    excel_bytes = await asyncio.to_thread(self.generate_diem_xlsx, result["data"])
                    filename = "diem_toan_bo.xlsx"
                    caption = "📄 Bảng điểm toàn bộ"
                else:
                    excel_bytes = await asyncio.to_thread(
                        self.generate_diem_xlsx, result["data"], export_type
                    )
                    hocky_name = (
                        result["data"].get("hocky_data", {}).get(export_type, {}).get("hocky_name", export_type)
                    )
                    filename = f"diem_{hocky_name}.xlsx"
                    caption = f"📄 Bảng điểm {hocky_name}"
                await self.telegram.send_document(
                    chat_id=chat_id, file=excel_bytes, filename=filename, caption=caption
                )
            except Exception as e:
                logger.error("Lỗi tạo file Excel: %s", e, exc_info=True)
                await self._safe_edit(chat_id, message_id, f"Lỗi tạo file Excel: {str(e)}")
            return

        # Xem điểm chi tiết học kỳ
        await self.telegram.answer_callback_query(callback_id, text="Đang tải điểm...")
        result = await self.handle_diem(user_id, hocky_key)
        if not result["success"]:
            await self._safe_edit(chat_id, message_id, result["message"], parse_mode="HTML")
            return
        try:
            await self.telegram.edit_message_text_rich(
                chat_id=chat_id,
                message_id=message_id,
                html=self.format_diem_detail_message(result["data"]),
                reply_markup=build_inline_keyboard([
                    [
                        make_inline_button("Xuất Excel", f"diem_export_{hocky_key}", tone="warning", emoji="📄"),
                        make_inline_button("Quay lại", "diem_back", tone="neutral"),
                    ]
                ]),
            )
        except TelegramAPIError as e:
            if "message is not modified" not in e.description.lower():
                raise

    # ==================== Data layer ====================

    async def handle_diem(self, telegram_user_id: int, hocky_key: Optional[str] = None) -> Dict[str, Any]:
        try:
            cache_key = f"diem:{telegram_user_id}"
            cached = await self.cache_manager.get(cache_key)
            if cached:
                processed = self._process_diem_data(cached.get("data"), hocky_key)
                processed["timestamp"] = cached.get("timestamp")
                return {"success": True, "message": "OK (cache)", "data": processed}

            token = await self._get_user_token(telegram_user_id)
            if not token:
                return {"success": False, "message": "Bạn chưa đăng nhập. Vui lòng /dangnhap.", "data": None}

            response = await self._call_diem_api(token)
            if isinstance(response, dict) and response.get("error"):
                return {
                    "success": False,
                    "message": self._format_api_error_message(response),
                    "data": None,
                    "error_type": "api_error",
                    "status_code": response.get("status_code"),
                }
            if response and isinstance(response, list):
                await self.cache_manager.set(cache_key, response, ttl=86400)
                processed = self._process_diem_data(response, hocky_key)
                cached_data = await self.cache_manager.get(cache_key)
                processed["timestamp"] = (
                    cached_data.get("timestamp") if cached_data else datetime.utcnow().isoformat()
                )
                return {"success": True, "message": "OK (mới)", "data": processed}
            return {
                "success": False,
                "message": "🚫 Lỗi: Không thể lấy dữ liệu điểm.",
                "data": response,
                "show_back_button": True,
            }
        except Exception as e:
            logger.error("Điểm error for user %s: %s", telegram_user_id, e)
            return {
                "success": False,
                "message": f"🚫 Lỗi: {str(e)}",
                "data": None,
                "show_back_button": True,
            }

    async def _call_diem_api(self, token: str) -> Optional[Any]:
        try:
            url = f"{self.config.HUTECH_API_BASE_URL}{self.config.HUTECH_DIEM_ENDPOINT}"
            headers = self.config.HUTECH_MOBILE_HEADERS.copy()
            headers["authorization"] = f"JWT {token}"
            async with aiohttp.ClientSession() as session:
                async with session.post(url, headers=headers, json={}) as response:
                    if response.status == 201:
                        return await response.json()
                    return {"error": True, "status_code": response.status, "message": await response.text()}
        except aiohttp.ClientError as e:
            return {"error": True, "message": f"Lỗi kết nối: {str(e)}"}
        except json.JSONDecodeError as e:
            return {"error": True, "message": f"Lỗi phân tích dữ liệu: {str(e)}"}
        except Exception as e:
            return {"error": True, "message": f"Lỗi không xác định: {str(e)}"}

    async def _get_user_token(self, telegram_user_id: int) -> Optional[str]:
        try:
            response_data = await self.db_manager.get_user_login_response(telegram_user_id)
            if not response_data:
                return None
            old_login_info = response_data.get("old_login_info")
            if isinstance(old_login_info, dict) and old_login_info.get("token"):
                return old_login_info["token"]
            return response_data.get("token")
        except Exception as e:
            logger.error("Error getting token for user %s: %s", telegram_user_id, e)
            return None

    # ==================== Process ====================

    @staticmethod
    def _process_diem_data(diem_data: List[Dict[str, Any]], hocky_key: Optional[str]) -> Dict[str, Any]:
        try:
            hocky_data: Dict[str, Any] = {}
            for hocky in diem_data:
                if "nam_hoc_hoc_ky" in hocky:
                    key = hocky["nam_hoc_hoc_ky"]
                    diem_chi_tiet = sorted(
                        hocky.get("diem_chi_tiet", []),
                        key=lambda x: x.get("ten_hp", ""),
                    )
                    hocky_data[key] = {
                        "hocky_name": hocky.get("nam_hoc_hoc_ky_name", ""),
                        "diem_chi_tiet": diem_chi_tiet,
                        "diem_tich_luy": hocky.get("diem_tich_luy", {}),
                    }
            if hocky_key and hocky_key in hocky_data:
                return {"selected_hocky": hocky_key, "hocky_data": {hocky_key: hocky_data[hocky_key]}}
            return {"selected_hocky": None, "hocky_data": hocky_data}
        except Exception as e:
            logger.error("Error processing điểm data: %s", e)
            return {"selected_hocky": None, "hocky_data": {}}

    @staticmethod
    def _format_api_error_message(error_data: Dict[str, Any]) -> str:
        status_code = error_data.get("status_code")
        error_message = error_data.get("message", "")
        if status_code == 422:
            try:
                err = json.loads(error_message)
                msg_text = err.get("errorMessage", "")
                if any(s in msg_text.lower() for s in ("khảo sát", "survey", "không đủ điều kiện", "not eligible")):
                    return (
                        "🚫 <b>Không thể xem điểm</b>\n\n"
                        "Bạn chưa hoàn thành các khảo sát sinh viên bắt buộc.\n\n"
                        "Để xem điểm, vui lòng:\n"
                        "1. Truy cập trang web sinhvien.hutech.edu.vn\n"
                        "2. Đăng nhập vào hệ thống\n"
                        "3. Hoàn thành đầy đủ các phiếu khảo sát tại mục \"Khảo sát sinh viên\"\n\n"
                        "Sau khi hoàn thành khảo sát, hãy thử lại lệnh /diem"
                    )
                reasons_msg = err.get("reasons", {}).get("message", msg_text)
                return f"🚫 Lỗi từ hệ thống: {escape_html(reasons_msg)}"
            except (json.JSONDecodeError, KeyError):
                if any(s in error_message.lower() for s in ("không đủ điều kiện", "not eligible")):
                    return (
                        "🚫 <b>Không thể xem điểm</b>\n\n"
                        "Bạn chưa hoàn thành các khảo sát sinh viên bắt buộc.\n\n"
                        "Vui lòng truy cập sinhvien.hutech.edu.vn để hoàn thành."
                    )
                return (
                    "🚫 <b>Không thể xem điểm</b>\n\n"
                    "Hệ thống báo lỗi: Sinh viên không đủ điều kiện để xem điểm."
                )
        if status_code == 401:
            return "🚫 <b>Lỗi xác thực</b>: Phiên đăng nhập đã hết hạn. Vui lòng /dangxuat và /dangnhap lại."
        if status_code == 403:
            return "🚫 <b>Lỗi quyền truy cập</b>: Bạn không có quyền truy cập chức năng này."
        if status_code == 404:
            return "🚫 <b>Không tìm thấy</b>: Không tìm thấy dữ liệu điểm."
        if status_code == 500:
            return "🚫 <b>Lỗi máy chủ</b>: Máy chủ đang gặp sự cố. Vui lòng thử lại sau."
        if isinstance(status_code, int) and status_code >= 500:
            return f"🚫 <b>Lỗi máy chủ</b>: {status_code}. Vui lòng thử lại sau."
        return f"🚫 <b>Lỗi API</b>: {escape_html(error_message or '')}"

    # ==================== Format (Rich Message) ====================

    def format_diem_menu_message(self, diem_data: Dict[str, Any]) -> str:
        hocky_data = diem_data.get("hocky_data", {})
        if not hocky_data:
            return join_blocks([
                section_heading("📊", "Bảng Điểm"),
                p("Không có dữ liệu điểm để hiển thị."),
            ])
        sorted_keys = sorted(hocky_data.keys(), reverse=True)
        recent_keys = sorted_keys[:3]
        rows: List[List[str]] = []
        for i, key in enumerate(recent_keys, 1):
            data = hocky_data[key]
            hocky_name = data.get("hocky_name", "N/A")
            tl = data.get("diem_tich_luy") or {}
            dtb = tl.get("diem_trung_binh_he_4", "—")
            tc_dat = tl.get("so_tin_chi_dat", "—")
            rows.append([str(i), hocky_name, str(dtb), str(tc_dat)])

        blocks: List[str] = [
            section_heading("📊", "Bảng Điểm Các Học Kỳ"),
            p("Chọn một học kỳ để xem chi tiết điểm hoặc xuất file Excel."),
            table(
                ["#", "Học kỳ", "Điểm TB (Hệ 4)", "Số TC Đạt"],
                rows,
                bordered=True,
                striped=True,
            ),
        ]
        ts_footer = footer_updated_at(diem_data.get("timestamp"))
        if ts_footer:
            blocks.append(ts_footer)
        return join_blocks(blocks)

    def format_older_hocky_menu_message(self, diem_data: Dict[str, Any]) -> str:
        hocky_data = diem_data.get("hocky_data", {})
        if not hocky_data:
            return join_blocks([
                section_heading("📊", "Các Học Kỳ Cũ Hơn"),
                p("Không có dữ liệu điểm để hiển thị."),
            ])
        sorted_keys = sorted(hocky_data.keys(), reverse=True)
        older = sorted_keys[3:]
        rows: List[List[str]] = []
        for i, key in enumerate(older, 1):
            data = hocky_data[key]
            hocky_name = data.get("hocky_name", "N/A")
            tl = data.get("diem_tich_luy") or {}
            dtb = tl.get("diem_trung_binh_he_4", "—")
            tc_dat = tl.get("so_tin_chi_dat", "—")
            rows.append([str(i), hocky_name, str(dtb), str(tc_dat)])

        blocks: List[str] = [
            section_heading("📊", "Các Học Kỳ Cũ Hơn"),
            p("Chọn một học kỳ để xem chi tiết điểm hoặc xuất file Excel."),
            table(
                ["#", "Học kỳ", "Điểm TB (Hệ 4)", "Số TC Đạt"],
                rows,
                bordered=True,
                striped=True,
            ),
        ]
        ts_footer = footer_updated_at(diem_data.get("timestamp"))
        if ts_footer:
            blocks.append(ts_footer)
        return join_blocks(blocks)

    def format_diem_detail_message(self, diem_data: Dict[str, Any]) -> str:
        hocky_data = diem_data.get("hocky_data", {})
        selected = diem_data.get("selected_hocky")
        if not hocky_data or not selected or selected not in hocky_data:
            return section_heading("📊", "Không có dữ liệu điểm chi tiết.")
        data = hocky_data[selected]
        hocky_name = data.get("hocky_name", "N/A")
        diem_chi_tiet = data.get("diem_chi_tiet", [])
        diem_tich_luy = data.get("diem_tich_luy", {})

        blocks: List[str] = [section_heading("📊", f"Điểm Chi Tiết - {hocky_name}")]
        if diem_tich_luy:
            summary_rows: List[List[str]] = [
                ["Điểm TB (Hệ 4)", str(diem_tich_luy.get("diem_trung_binh_he_4", "—"))],
                ["Điểm TB Tích Lũy (Hệ 4)", str(diem_tich_luy.get("diem_trung_binh_tich_luy_he_4", "—"))],
                ["Số TC Đạt", str(diem_tich_luy.get("so_tin_chi_dat", "—"))],
                ["Tổng TC Tích Lũy", str(diem_tich_luy.get("so_tin_chi_tich_luy", "—"))],
            ]
            blocks.append(h2("Tổng Kết Học Kỳ"))
            blocks.append(table(
                ["Chỉ số", "Giá trị"],
                summary_rows,
                bordered=True,
                striped=True,
            ))

        if diem_chi_tiet:
            blocks.append(hr())
            blocks.append(h2("Điểm Môn Học"))
            # Tạo bảng
            headers = ["STT", "Mã HP", "Tên học phần", "STC", "KT1", "KT2", "Thi", "Điểm 10", "Điểm 4", "Điểm chữ"]
            rows: List[List[str]] = []
            for i, mon in enumerate(diem_chi_tiet, 1):
                kt1 = mon.get("diem_kiem_tra_1", "") or ""
                kt2 = mon.get("diem_kiem_tra_2", "") or ""
                thi = mon.get("diem_thi", "") or ""
                rows.append([
                    str(i),
                    str(mon.get("ma_hp", "N/A")),
                    str(mon.get("ten_hp", "N/A")),
                    str(mon.get("stc", "N/A")),
                    str(kt1),
                    str(kt2),
                    str(thi),
                    str(mon.get("diem_he_10", "N/A")),
                    str(mon.get("diem_he_4", "N/A")),
                    str(mon.get("diem_chu", "N/A")),
                ])
            blocks.append(table(headers, rows, bordered=True, striped=True))
        else:
            blocks.append(p("Không có điểm chi tiết trong học kỳ này."))

        ts_footer = footer_updated_at(diem_data.get("timestamp"))
        if ts_footer:
            blocks.append(ts_footer)
        return join_blocks(blocks)

    # ==================== Keyboards & lists ====================

    def get_hocky_list(self, diem_data: Dict[str, Any]) -> List[Dict[str, str]]:
        hocky_data = diem_data.get("hocky_data", {})
        if not hocky_data:
            return []
        sorted_keys = sorted(hocky_data.keys(), reverse=True)
        result: List[Dict[str, str]] = []
        for i, key in enumerate(sorted_keys[:3]):
            data = hocky_data[key]
            result.append({
                "key": key,
                "name": data.get("hocky_name", ""),
                "display": str(i + 1),
            })
        if len(sorted_keys) > 3:
            result.append({"key": "more", "name": "Xem thêm học kỳ cũ hơn", "display": "4"})
        return result

    def get_older_hocky_list(self, diem_data: Dict[str, Any]) -> List[Dict[str, str]]:
        hocky_data = diem_data.get("hocky_data", {})
        if not hocky_data:
            return []
        sorted_keys = sorted(hocky_data.keys(), reverse=True)
        older = sorted_keys[3:]
        return [
            {"key": key, "name": hocky_data[key].get("hocky_name", ""), "display": str(i + 1)}
            for i, key in enumerate(older)
        ]

    def create_main_diem_keyboard(self, diem_data: Dict[str, Any]) -> Dict[str, Any]:
        rows: List[List[Dict[str, Any]]] = []
        for h in self.get_hocky_list(diem_data):
            rows.append([make_inline_button(h["name"], f"diem_{h['key']}", tone=None)])
        rows.append([make_inline_button("Xuất Excel toàn bộ", "diem_export_all", tone="warning", emoji="📄")])
        return build_inline_keyboard(rows)

    # ==================== XLSX ====================

    def generate_diem_xlsx(self, diem_data: Dict[str, Any], hocky_key: Optional[str] = None) -> bytes:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            title_font = Font(name="Arial", size=16, bold=True)
            header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
            cell_font = Font(name="Arial", size=11)
            tich_luy_font = Font(name="Arial", size=11, bold=True)
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left = Alignment(horizontal="left", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )
            hocky_data = diem_data.get("hocky_data", {})

            if hocky_key and hocky_key in hocky_data:
                data = hocky_data[hocky_key]
                hocky_name = data.get("hocky_name", "")
                ws.title = hocky_name
                self._write_hocky_to_sheet(ws, hocky_name, data, title_font, header_font, cell_font, tich_luy_font, header_fill, center, left, border)
            else:
                ws.title = "Điểm Toàn Bộ"
                sorted_keys = sorted(hocky_data.keys())
                current_row = 1
                for key in sorted_keys:
                    data = hocky_data[key]
                    hocky_name = data.get("hocky_name", "")
                    current_row = self._write_hocky_to_sheet(
                        ws, hocky_name, data, title_font, header_font, cell_font, tich_luy_font,
                        header_fill, center, left, border, start_row=current_row,
                    )
                    current_row += 2
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()
        except Exception as e:
            logger.error("Error generating điểm XLSX: %s", e, exc_info=True)
            raise

    def _write_hocky_to_sheet(self, ws, hocky_name, data, title_font, header_font, cell_font,
                               tich_luy_font, header_fill, center, left, border, start_row=1) -> int:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
        cell = ws.cell(row=start_row, column=1, value=f"BẢNG ĐIỂM HỌC KỲ: {hocky_name.upper()}")
        cell.font = title_font
        cell.alignment = center

        headers = ["STT", "Mã HP", "Tên học phần", "STC", "KT1", "KT2", "Thi", "Điểm 10", "Điểm 4", "Điểm chữ"]
        header_row = start_row + 1
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=header_row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border

        diem_chi_tiet = data.get("diem_chi_tiet", [])
        current_row = header_row + 1
        for i, mon in enumerate(diem_chi_tiet, 1):
            values = [
                i,
                mon.get("ma_hp", ""),
                mon.get("ten_hp", ""),
                mon.get("stc", ""),
                mon.get("diem_kiem_tra_1", ""),
                mon.get("diem_kiem_tra_2", ""),
                mon.get("diem_thi", ""),
                mon.get("diem_he_10", ""),
                mon.get("diem_he_4", ""),
                mon.get("diem_chu", ""),
            ]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=current_row, column=col, value=val)
                c.font = cell_font
                c.border = border
                c.alignment = center if col in (1, 4, 5, 6, 7, 8, 9, 10) else left
            current_row += 1

        diem_tich_luy = data.get("diem_tich_luy", {})
        if diem_tich_luy:
            tich_luy_data = [
                ("Điểm TB học kỳ (hệ 4)", diem_tich_luy.get("diem_trung_binh_he_4", "")),
                ("Điểm TB tích lũy (hệ 4)", diem_tich_luy.get("diem_trung_binh_tich_luy_he_4", "")),
                ("Số TC đạt", diem_tich_luy.get("so_tin_chi_dat", "")),
                ("Tổng TC tích lũy", diem_tich_luy.get("so_tin_chi_tich_luy", "")),
            ]
            for i, (label, value) in enumerate(tich_luy_data):
                ws.merge_cells(start_row=current_row + i, start_column=1, end_row=current_row + i, end_column=3)
                lbl = ws.cell(row=current_row + i, column=1, value=label)
                lbl.font = tich_luy_font
                lbl.alignment = left
                val = ws.cell(row=current_row + i, column=4, value=value)
                val.font = tich_luy_font
                val.alignment = center

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 5
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 8
        ws.column_dimensions["G"].width = 8
        ws.column_dimensions["H"].width = 10
        ws.column_dimensions["I"].width = 10
        ws.column_dimensions["J"].width = 10

        return current_row + len(tich_luy_data) if diem_tich_luy else current_row

    # ==================== Utils ====================

    async def _safe_edit(self, chat_id: int, message_id: int, text: str, **kwargs) -> None:
        try:
            await self.telegram.edit_message_text_plain(
                chat_id=chat_id, message_id=message_id, text=text, **kwargs
            )
        except TelegramAPIError as e:
            if "message is not modified" not in e.description.lower():
                logger.debug("safe_edit failed: %s", e.description)
